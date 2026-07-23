import io
import json
from datetime import datetime
from urllib.parse import quote

from flask import Blueprint, Response, jsonify, request

from . import service
from .schema import (
    normalize_audit_export,
    normalize_audit_list,
    normalize_call_statistics,
)

audits_bp = Blueprint("audits", __name__)


@audits_bp.post("/list")
def list_audits():
    payload = request.get_json(silent=True) or {}
    query = normalize_audit_list(payload)
    data = service.list_audit_logs(query)

    return jsonify({
        "msg_id": f"{payload.get('msg_id', '')}_Resp".replace("_Resp_Resp", "_Resp"),
        "serial": payload.get("serial", ""),
        "context": payload.get("context", ""),
        "http_status_code": 200,
        "is_success": True,
        "content": {
            "list": data.get("list", []),
            "total": data.get("total", 0),
            "page": data.get("page", 1),
            "page_size": data.get("page_size", 100),
        },
    })


@audits_bp.post("/export")
def export_audits():
    payload = request.get_json(silent=True) or {}
    query = normalize_audit_export(payload)
    records = service.export_audit_logs(query)
    fmt = query.get("format", "json")

    if fmt == "excel":
        return _export_excel(records, query)
    return _export_json(records, query)


@audits_bp.get("/call-statistics")
def call_statistics():
    query, error = normalize_call_statistics(request.args)
    if error:
        return jsonify({
            "is_success": False,
            "http_status_code": 400,
            "msg": error,
        }), 400

    return jsonify(service.get_call_statistics(query["time_range"]))


def _export_filename(extension: str, time_range: str):
    range_name = time_range if time_range in {"1h", "1d", "7d", "30d"} else "全部"
    return f"{datetime.now():%Y%m%d}_审计日志_{range_name}.{extension}"


def _download_headers(filename: str):
    ascii_fallback = f"audit_logs_{datetime.now():%Y%m%d}.{filename.rsplit('.', 1)[-1]}"
    return {
        "Content-Disposition": (
            f"attachment; filename*=UTF-8''{quote(filename)}; "
            f"filename={ascii_fallback}"
        )
    }


def _export_json(records: list, query: dict):
    json_str = json.dumps(records, ensure_ascii=False, indent=2, default=str)
    filename = _export_filename("json", query.get("time_range", "all"))
    return Response(
        json_str,
        mimetype="application/json",
        headers=_download_headers(filename),
    )


def _export_excel(records: list, query: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "操作审计日志"

    headers = [
        "ID",
        "操作类型",
        "操作人",
        "操作人IP",
        "操作对象",
        "对象名称",
        "HTTP状态码",
        "是否成功",
        "错误信息",
        "创建时间",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, record in enumerate(records, 2):
        values = [
            record.get("id"),
            record.get("operation_type"),
            record.get("operator"),
            record.get("operator_ip"),
            record.get("target_type"),
            record.get("target_name"),
            record.get("http_status_code"),
            "成功" if record.get("is_success") else "失败",
            record.get("error_message", ""),
            record.get("created_at"),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="center")

    for col, width in enumerate([6, 14, 12, 16, 12, 28, 14, 10, 40, 22], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = _export_filename("xlsx", query.get("time_range", "all"))

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_download_headers(filename),
    )
